# -*- coding: utf-8 -*-
"""
Created on Wed Apr 21 17:38:46 2021

@author: James Ang
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl.styles as sty
from openpyxl import Workbook ,load_workbook
import json
#Load xlsx
wb=load_workbook('E://Summary of Economic Data.xlsx')
#Read multiple sheets
for ws in wb: 
    c = ws.max_column
    r = ws.max_row
    list_key=[]
    jsonLine=[]
    datalines=[]
    dict_v={}
    dict_v["name"]=ws.title
#Set the year as the key
    for col in range(2,c+1):
        list_key.append(ws.cell(row=1,column=col).value)
#Read annual data
    for col in range(2,c+1):
        dataline=[]
        for row in range(2,r+1):
            data={}       
            data["year"]=ws.cell(row=row,column=1).value
            data["value"]=ws.cell(row=row,column=col).value
            dataline.append(data)
        dict_c={}
        dict_c["Country"]=list_key[col-2]
        dict_c["data"]=dataline
        datalines.append(dict_c)
    dict_v["data"]=datalines
    
#Save into json format    
	json.dump(dict_v,open(ws.title+".json","w",encoding='utf-8'),ensure_ascii=False)


